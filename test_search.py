import argparse
from urllib.parse import urlencode, urljoin

import openpyxl
import requests

parser = argparse.ArgumentParser(description='test nominatim direct and reverse geocoding')

parser.add_argument("file_name", type=str, help="file name with data for tests")

args = parser.parse_args()

BASE_COLUMNS = {
    'input_value': 1,
    'expected_result': 2,
    'actual_response': 3,
    'test_result': 4,
}

data_file = args.file_name

try:
    wb = openpyxl.load_workbook(data_file)
    ws = wb.active
except FileNotFoundError:
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, j in BASE_COLUMNS.items():
        ws.cell(row=1, column=j, value=i)
    wb.save(data_file)

failed_code = 'failed({})'
ok_code = 'ok'
error_code = 'error'
bad_response_code = 'bad response - {}'


def nominatim_request(query, params):
    base_url = 'https://nominatim.openstreetmap.org/'
    base_params = {'format': 'json', 'accept-language': 'en-US'}
    base_params.update(params)
    url = urljoin(base_url, f'{query}{urlencode(base_params)}')
    print(url)
    result = None
    try:
        result = requests.get(url)
        return result.json()
    except Exception as e:
        if result:
            return result.status_code
        return str(e)


def geocoding(search_query):
    return nominatim_request('search?', {'q': search_query})


def reverse_geocoding(lat, lon):
    return nominatim_request('reverse?', {'lat': lat, 'lon': lon})


def save_value(result, row, col):
    ws.cell(row=row, column=col, value=result)
    wb.save(data_file)


def str_to_float_coordinates(coordinates):
    return [float(i) for i in coordinates]


def is_coordinates(value):
    return value.replace(',', '').replace('.', '').replace(' ', '').isdigit()


def test_reverse_geocoding(value, expected_result, current_row):
    lat, lon = value.replace(' ', '').split(',')
    reverse_geocoding_result = reverse_geocoding(lat, lon)

    print(reverse_geocoding_result)
    print(f'{"-"*30}')

    if type(reverse_geocoding_result) in (str, int):
        save_value('', current_row, BASE_COLUMNS['actual_response'])
        save_value(
            failed_code.format(bad_response_code.format(reverse_geocoding_result)),
            current_row,
            BASE_COLUMNS['test_result']
        )
        return

    if reverse_geocoding_result.get(error_code):
        save_value(error_code, current_row, BASE_COLUMNS['actual_response'])
        result = ok_code if expected_result == error_code else failed_code.format(error_code)
        save_value(result, current_row, BASE_COLUMNS['test_result'])
        return

    display_name = reverse_geocoding_result['display_name']
    save_value(display_name, current_row, BASE_COLUMNS['actual_response'])
    result = ok_code if expected_result is None or expected_result in display_name else error_code
    save_value(result, current_row, BASE_COLUMNS['test_result'])


def test_geocoding(address, expected_result, current_row):
    geocoding_result = geocoding(address)

    print(geocoding_result)
    print(f'{"-"*30}')

    if type(geocoding_result) in (str, int):
        save_value('', current_row, BASE_COLUMNS['actual_response'])
        save_value(
            failed_code.format(bad_response_code.format(geocoding_result)),
            current_row,
            BASE_COLUMNS['test_result']
        )
        return

    if expected_result is None:
        result = ok_code if geocoding_result else failed_code.format(error_code)
        save_value(result, current_row, BASE_COLUMNS['test_result'])
        save_value(str(geocoding_result), current_row, BASE_COLUMNS['actual_response'])
        return

    failed = failed_code.format(error_code)
    test_result = ok_code if expected_result == error_code else failed
    actual_response = geocoding_result

    if expected_result != error_code:
        lat, lon = str_to_int_coordinates(expected_result.replace(' ', '').split(','))

        for i in geocoding_result:
            b_box = str_to_int_coordinates(i['boundingbox'])
            test_result = ok_code if b_box[0] <= lat <= b_box[1] and b_box[2] <= lon <= b_box[3] else failed
            if test_result:
                actual_response = i
                break

    save_value(test_result, current_row, BASE_COLUMNS['test_result'])
    save_value(str(actual_response), current_row, BASE_COLUMNS['actual_response'])


row = 2

while True:
    test_value = ws.cell(row=row, column=BASE_COLUMNS['input_value']).value
    test_expected_result = ws.cell(row=row, column=BASE_COLUMNS['expected_result']).value
    if test_value is None:
        break

    print(test_value)
    print(test_expected_result)

    if is_coordinates(test_value):
        test_reverse_geocoding(test_value, test_expected_result, row)
    else:
        test_geocoding(test_value, test_expected_result, row)

    row += 1
